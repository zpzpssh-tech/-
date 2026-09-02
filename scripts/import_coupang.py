"""
쿠팡 계정별 엑셀 → 매출·원가·비용 변환기
----------------------------------------
쓰는 법:
  1) data/원본/쿠팡/ 폴더에 계정별 엑셀을 넣습니다.
     파일 이름이 곧 계정 이름이 됩니다. 예: 휴책.xlsx, 유큐어.xlsx, 올투게더나우.xlsx
  2) python3 scripts/import_coupang.py
  3) python3 scripts/build.py

엑셀 구조 (쿠팡 판매자센터에서 받은 그대로):
  [원가]              옵션명 | 상품명 | 원가
  [YYYYMM월]          날짜 | 방문자 | 조회 | 장바구니 | 주문 | 판매량 | 구매전환율 | 매출(원)
                      → 하루 합계만 있고 상품 구분이 없습니다.
  [YYYYMMDD]          옵션명 | 상품명 | 매출(원) | 주문 | 판매량 | 방문자 | 조회 | 장바구니 | 구매전환율
                      → 그날 상품별 실적입니다.
  [월별광고/N월광고]   날짜 | 배송유형(SELLER/ROCKETGROWTH) | 청구금액(+부가가치세)
  [월별그로스부대비용] 날짜 | 판매자 할인쿠폰 | 입출고비&배송비 | 수수료
     → 로켓그로스로 파는 계정에만 있습니다. 판매자배송만 하는 계정은 이 시트가 없습니다.

원가 붙이는 순서 (위에서부터 먼저 맞는 것을 씁니다):
  1. 옵션명이 정확히 같을 때
  2. 상품명이 같고 옵션 꼬리표(', 1개, free' 같은 부분)도 같을 때
  3. 상품명이 같고 그 상품의 원가가 하나뿐일 때
  4. 상품명이 같을 때 그 상품 원가들의 평균
  못 찾은 건은 data/쿠팡_원가미매칭.csv 에 적어 둡니다.

상품별 실적이 없는 달(월별 시트만 있는 달)은 상품별 실적이 있는 달에서 구한
'매출 대비 원가 비율'을 적용해 원가를 추정합니다. 추정한 달은 표시해 둡니다.
"""
import csv
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("[오류] openpyxl이 없습니다. 'pip install openpyxl' 을 먼저 실행해 주세요.")

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "원본" / "쿠팡"
DATA = ROOT / "data"
CHANNEL = "쿠팡"


def num(v):
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", "").replace("₩", "").strip())
    except ValueError:
        return 0.0


def as_date(v):
    """여러 형태의 날짜를 YYYY-MM-DD 로 통일합니다."""
    s = str(v).strip()
    if re.fullmatch(r"\d{8}", s):
        return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    s = s[:10].replace(".", "-").replace("/", "-")
    return s if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s) else ""


def tail(opt):
    """옵션명에서 뒤쪽 꼬리표만 뽑습니다. '상품명, 1개, free' → '1개,free'"""
    parts = [p.strip() for p in str(opt).split(",")]
    return ",".join(parts[1:]) if len(parts) > 1 else ""


def header_of(ws):
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    return [str(h).strip() if h is not None else "" for h in (first or [])]


def read_account(path):
    account = path.stem
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    # ── 원가 시트 ──
    if "원가" not in wb.sheetnames:
        wb.close()
        sys.exit(f"[오류] {path.name}에 '원가' 시트가 없습니다. 시트: {', '.join(wb.sheetnames)}")
    ws = wb["원가"]; ws.reset_dimensions()
    it = ws.iter_rows(values_only=True); next(it, None)
    by_opt, by_prod = {}, defaultdict(dict)
    for r in it:
        r = list(r) + [None, None, None]
        if r[0] is None:
            continue
        opt, prod, c = str(r[0]).strip(), str(r[1] or "").strip(), num(r[2])
        by_opt[opt] = c
        if prod:
            by_prod[prod][tail(opt)] = c

    def find_cost(opt, prod):
        if opt in by_opt:
            return by_opt[opt], "옵션명 일치"
        d = by_prod.get(prod)
        if d:
            if tail(opt) in d:
                return d[tail(opt)], "상품명+꼬리표"
            if len(d) == 1:
                return next(iter(d.values())), "상품명 원가 1개"
            return sum(d.values()) / len(d), "상품명 평균"
        return None, "못 찾음"

    # ── 판매 시트 ──
    daily = defaultdict(lambda: {"ord": 0, "qty": 0, "rev": 0, "cogs": 0, "unk": 0,
                                 "visit": 0, "view": 0, "cart": 0, "est": 0})
    options = defaultdict(lambda: {"rev": 0, "ord": 0, "qty": 0, "cogs": 0})
    prod_days = set()          # 상품별 실적이 있는 날짜
    tiers, missing = Counter(), Counter()

    for sh in wb.sheetnames:
        is_month = bool(re.fullmatch(r"\d{6}월", sh))
        is_day = bool(re.fullmatch(r"\d{8}", sh))
        if not (is_month or is_day):
            continue
        ws = wb[sh]; ws.reset_dimensions()
        hdr = header_of(ws)
        if "매출(원)" not in hdr:
            continue
        I = {c: hdr.index(c) for c in hdr if c}
        width = len(hdr)
        it = ws.iter_rows(min_row=2, values_only=True)
        if is_day:
            date = as_date(sh)
            prod_days.add(date)
            for r in it:
                r = list(r) + [None] * width
                if r[0] is None:
                    continue
                rev, orders, qty = num(r[I["매출(원)"]]), num(r[I["주문"]]), num(r[I["판매량"]])
                opt = str(r[I["옵션명"]] or "").strip()
                prod = str(r[I.get("상품명", I["옵션명"])] or "").strip()
                d = daily[date]
                d["rev"] += rev; d["ord"] += orders; d["qty"] += qty
                for k, col in (("visit", "방문자"), ("view", "조회"), ("cart", "장바구니")):
                    if col in I:
                        d[k] += num(r[I[col]])
                if qty <= 0:
                    continue
                unit, tier = find_cost(opt, prod)
                tiers[tier] += qty
                if unit is None:
                    d["unk"] += qty
                    missing[(prod[:60], tail(opt)[:40])] += qty
                else:
                    d["cogs"] += unit * qty
                o = options[(date[:7], prod, opt)]
                o["rev"] += rev; o["ord"] += orders; o["qty"] += qty
                o["cogs"] += (unit or 0) * qty
        else:
            for r in it:
                r = list(r) + [None] * width
                date = as_date(r[0]) if r[0] is not None else ""
                if not date:
                    continue
                d = daily[date]
                d["rev"] += num(r[I["매출(원)"]]); d["ord"] += num(r[I["주문"]]); d["qty"] += num(r[I["판매량"]])
                for k, col in (("visit", "방문자"), ("view", "조회"), ("cart", "장바구니")):
                    if col in I:
                        d[k] += num(r[I[col]])

    # 상품별 실적이 없는 날은 있는 날의 원가율로 추정합니다.
    known_rev = sum(daily[d]["rev"] for d in prod_days)
    known_cogs = sum(daily[d]["cogs"] for d in prod_days)
    rate = known_cogs / known_rev if known_rev else 0
    for date, d in daily.items():
        if date not in prod_days and d["rev"]:
            d["cogs"] = d["rev"] * rate
            d["est"] = 1

    # ── 광고비 ──
    ads = defaultdict(float)       # (날짜, 배송유형)
    ad_month_only = defaultdict(float)
    covered_months = set()
    for sh in wb.sheetnames:
        if "광고" not in sh or sh == "월별광고":
            continue
        ws = wb[sh]; ws.reset_dimensions()
        for r in ws.iter_rows(min_row=2, values_only=True):
            r = list(r) + [None, None, None]
            date = as_date(r[0]) if r[0] is not None else ""
            if not date or r[2] is None:
                continue
            ads[(date, str(r[1] or "").strip())] += num(r[2])
            covered_months.add(date[:7])
    if "월별광고" in wb.sheetnames:
        ws = wb["월별광고"]; ws.reset_dimensions()
        for r in ws.iter_rows(min_row=2, values_only=True):
            r = list(r) + [None, None, None]
            date = as_date(r[0]) if r[0] is not None else ""
            if not date or r[2] is None or date[:7] in covered_months:
                continue
            ad_month_only[(date[:7], str(r[1] or "").strip())] += num(r[2])

    # ── 로켓그로스 부대비용 ──
    fees = defaultdict(lambda: {"쿠폰": 0.0, "입출고비": 0.0, "수수료": 0.0})
    fee_month_only = defaultdict(lambda: {"쿠폰": 0.0, "입출고비": 0.0, "수수료": 0.0})
    fee_months = set()
    for sh in wb.sheetnames:
        if "그로스부대비용" not in sh or sh == "월별그로스부대비용":
            continue
        ws = wb[sh]; ws.reset_dimensions()
        for r in ws.iter_rows(min_row=2, values_only=True):
            r = list(r) + [None] * 4
            date = as_date(r[0]) if r[0] is not None else ""
            if not date:
                continue
            f = fees[date]
            f["쿠폰"] += num(r[1]); f["입출고비"] += num(r[2]); f["수수료"] += num(r[3])
            fee_months.add(date[:7])
    if "월별그로스부대비용" in wb.sheetnames:
        ws = wb["월별그로스부대비용"]; ws.reset_dimensions()
        for r in ws.iter_rows(min_row=2, values_only=True):
            r = list(r) + [None] * 4
            s0 = str(r[0] or "").strip()
            m = as_date(r[0])[:7] if as_date(r[0]) else (s0[:7] if re.fullmatch(r"\d{4}-\d{2}", s0[:7]) else "")
            if not m or m in fee_months:
                continue
            f = fee_month_only[m]
            f["쿠폰"] += num(r[1]); f["입출고비"] += num(r[2]); f["수수료"] += num(r[3])

    wb.close()
    return dict(account=account, daily=daily, options=options, prod_days=prod_days,
                cost_rate=rate, tiers=tiers, missing=missing,
                ads=ads, ad_month_only=ad_month_only,
                fees=fees, fee_month_only=fee_month_only,
                has_growth=bool(fees or fee_month_only))


def main():
    files = sorted(p for p in SRC.glob("*.xlsx") if not p.name.startswith("~$"))
    if not files:
        sys.exit(f"[오류] 엑셀 파일이 없습니다. 계정별 엑셀을 여기에 넣어 주세요:\n       {SRC}\n"
                 f"       파일 이름이 계정 이름이 됩니다. 예: 휴책.xlsx")

    accounts = [read_account(p) for p in files]

    # ── 일별 매출·원가 ──
    with open(DATA / "쿠팡_일별.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["날짜", "채널", "계정", "주문", "판매량", "총매출", "원가", "원가추정",
                    "원가미상수량", "방문자", "조회", "장바구니"])
        for a in accounts:
            for date in sorted(a["daily"]):
                d = a["daily"][date]
                if not (d["rev"] or d["ord"]):
                    continue
                w.writerow([date, CHANNEL, a["account"], int(d["ord"]), int(d["qty"]),
                            round(d["rev"]), round(d["cogs"]), d["est"], int(d["unk"]),
                            int(d["visit"]), int(d["view"]), int(d["cart"])])

    # ── 옵션별 월 판매량 ──
    with open(DATA / "쿠팡_옵션.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["월", "계정", "상품명", "옵션명", "매출", "주문", "판매량", "원가합계"])
        for a in accounts:
            for (m, prod, opt), o in sorted(a["options"].items()):
                if o["qty"] <= 0:
                    continue
                w.writerow([m, a["account"], prod, opt, round(o["rev"]), int(o["ord"]),
                            int(o["qty"]), round(o["cogs"])])

    # ── 비용 (광고비·수수료·입출고비·쿠폰) ──
    with open(DATA / "쿠팡_비용.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["기간", "단위", "계정", "항목", "금액"])
        for a in accounts:
            for (date, kind), v in sorted(a["ads"].items()):
                if v:
                    w.writerow([date, "일", a["account"], f"광고비({kind or '기타'})", round(v)])
            for (m, kind), v in sorted(a["ad_month_only"].items()):
                if v:
                    w.writerow([m, "월", a["account"], f"광고비({kind or '기타'})", round(v)])
            for date, fr in sorted(a["fees"].items()):
                for k, v in fr.items():
                    if v:
                        w.writerow([date, "일", a["account"], k, round(v)])
            for m, fr in sorted(a["fee_month_only"].items()):
                for k, v in fr.items():
                    if v:
                        w.writerow([m, "월", a["account"], k, round(v)])

    # ── 원가 미매칭 ──
    with open(DATA / "쿠팡_원가미매칭.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["수량", "계정", "상품명", "옵션꼬리표", "원가"])
        for a in accounts:
            for (prod, t), q in a["missing"].most_common():
                w.writerow([int(q), a["account"], prod, t, ""])

    # ── 보고 ──
    print(f"쿠팡 계정 {len(accounts)}개 읽음: {', '.join(a['account'] for a in accounts)}\n")
    print(f"{'계정':<12}{'매출(총액)':>16}{'주문':>9}{'원가':>15}{'원가율':>8}{'반영률':>8}  판매방식")
    for a in accounts:
        rev = sum(d["rev"] for d in a["daily"].values())
        orders = sum(d["ord"] for d in a["daily"].values())
        cogs = sum(d["cogs"] for d in a["daily"].values())
        tot = sum(a["tiers"].values()) or 1
        cover = (tot - a["tiers"]["못 찾음"]) / tot * 100
        way = "로켓그로스" if a["has_growth"] else "판매자배송"
        print(f"{a['account']:<12}{rev:>16,.0f}{orders:>9,.0f}{cogs:>15,.0f}"
              f"{cogs / rev * 100 if rev else 0:>7.1f}%{cover:>7.1f}%  {way}")
    est_days = sum(1 for a in accounts for d in a["daily"].values() if d["est"])
    if est_days:
        print(f"\n※ 상품별 실적이 없어 원가를 비율로 추정한 날: {est_days}일")
        print("   (쿠팡 월별 시트에는 상품 구분이 없어, 일별 시트가 있는 달의 원가율을 적용했습니다)")
    miss = sum(sum(a["missing"].values()) for a in accounts)
    if miss:
        print(f"\n※ 원가를 못 찾은 수량 {miss:,.0f}개 → data/쿠팡_원가미매칭.csv")
    print("\n주의: 쿠팡 매출은 수수료를 빼기 전 총액입니다. 네이버 정산예정금과 기준이 다릅니다.")
    print("      대시보드에서는 수수료를 따로 빼서 비교합니다.")


if __name__ == "__main__":
    main()
