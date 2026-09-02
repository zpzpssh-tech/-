"""
고정지출·마케팅비·광고비 엑셀 → costs.csv 변환기
------------------------------------------------
쓰는 법:
  1) data/원본/비용/ 폴더에 아래 두 종류 파일을 넣습니다.
     · 고정지출/마케팅비 시트  (파일명에 '고정' 또는 '마케팅' 포함)
     · 네이버 검색광고 리포트  (파일명에 '광고' 포함)
  2) python3 scripts/import_expenses.py
  3) python3 scripts/build.py

고정지출 엑셀 형식:
  '올투게더나우 고정지출(항목) 1~8월' 같은 제목 줄 아래에 [항목 | 금액]이 이어집니다.
  제목에 적힌 기간(1~8월 / 9월부터)에 해당하는 달마다 그 금액을 적용합니다.
  오른쪽에 '마케팅비 …' 제목이 있으면 그 아래 항목도 같은 방식으로 읽습니다.

광고 리포트 형식 (두 가지를 모두 지원):
  · 월별 시트: '일별' 열에 날짜가 있는 형태
  · 일별 시트: 시트 이름이 20260701 처럼 날짜인 형태
  두 경우 모두 '총비용'을 그 달의 네이버 광고비로 더합니다.
"""
import csv
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("[오류] openpyxl이 없습니다. 'pip install openpyxl' 을 먼저 실행해 주세요.")

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "원본" / "비용"
DATA = ROOT / "data"
AD_ACCOUNT = "네이버"          # 검색광고비를 붙일 계정

TITLE = re.compile(r"(고정지출|마케팅비)")
# '1~8월', '9월부터', '1월~8월', '9월 부터' 등에서 적용 기간을 읽습니다.
RANGE_TO = re.compile(r"(\d{1,2})\s*(?:월)?\s*[~\-]\s*(\d{1,2})\s*월")
RANGE_FROM = re.compile(r"(\d{1,2})\s*월\s*부터")


def months_of(title: str, year: int):
    """제목에 적힌 기간을 'YYYY-MM' 목록으로 바꿉니다. 못 읽으면 None."""
    m = RANGE_TO.search(title)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        return [f"{year}-{x:02d}" for x in range(a, b + 1)]
    m = RANGE_FROM.search(title)
    if m:
        return [f"{year}-{x:02d}" for x in range(int(m.group(1)), 13)]
    return None


def num(v):
    if v is None:
        return None
    s = str(v).replace(",", "").replace("₩", "").strip()
    if not s:
        return None
    try:
        return int(round(float(s)))
    except ValueError:
        return None


def read_fixed(path, year):
    """고정지출/마케팅비 시트를 읽어 [(월, 구분, 항목, 금액)] 로 만듭니다."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out, blocks = [], []
    for ws in wb.worksheets:
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        # 제목 칸(예: '올투게더나우 고정지출(항목) 1~8월')을 찾습니다.
        heads = []
        for i, row in enumerate(grid):
            for j, v in enumerate(row):
                if v is None:
                    continue
                t = str(v).strip()
                if TITLE.search(t) and months_of(t, year):
                    heads.append((i, j, t))
        for i, j, title in heads:
            kind = "고정지출" if "고정지출" in title else "마케팅비"
            mons = months_of(title, year)
            items = []
            for r in range(i + 1, len(grid)):
                row = grid[r]
                name = row[j] if j < len(row) else None
                amt = num(row[j + 1]) if j + 1 < len(row) else None
                if name is None or str(name).strip() == "":
                    # 빈 줄이 나오면 그 블록은 끝
                    if items:
                        break
                    continue
                nm = str(name).strip()
                if TITLE.search(nm) and months_of(nm, year):
                    break
                if nm in ("합계", "총계", "소계"):
                    break
                if amt is None:
                    continue
                items.append((nm, amt))
            if items:
                blocks.append((title, kind, mons, items))
                for m in mons:
                    for nm, amt in items:
                        out.append((m, kind, nm, amt))
    wb.close()
    return out, blocks


def read_ads(path):
    """네이버 검색광고 리포트를 읽어 {월: 광고비} 로 만듭니다."""
    wb = openpyxl.load_workbook(path, data_only=True)
    by_month = defaultdict(float)
    conv_cnt = defaultdict(float)      # 광고 전환수
    conv_rev = defaultdict(float)      # 광고 전환매출 (ROAS 계산용)
    days = set()
    for sh in wb.sheetnames:
        ws = wb[sh]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            continue
        hdr = [str(h).strip() if h else "" for h in first]
        if "총비용" not in hdr:
            continue
        ci = hdr.index("총비용")
        ni = hdr.index("총 전환수") if "총 전환수" in hdr else None
        vi = next((hdr.index(c) for c in ("총 전환매출액(원)", "총 전환매출액") if c in hdr), None)
        if "일별" in hdr:                                  # 날짜가 열에 있는 시트
            di = hdr.index("일별")
            for r in ws.iter_rows(min_row=2, values_only=True):
                if di >= len(r) or r[di] is None:
                    continue
                d = str(r[di]).strip().rstrip(".").replace(".", "-").replace("/", "-")[:10]
                if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", d):
                    continue
                v = num(r[ci]) if ci < len(r) else None
                if v:
                    by_month[d[:7]] += v
                    days.add(d)
                if ni is not None and ni < len(r):
                    conv_cnt[d[:7]] += num(r[ni]) or 0
                if vi is not None and vi < len(r):
                    conv_rev[d[:7]] += num(r[vi]) or 0
        elif re.fullmatch(r"\d{8}", sh):                   # 시트 이름이 날짜인 시트
            d = f"{sh[:4]}-{sh[4:6]}-{sh[6:]}"
            for r in ws.iter_rows(min_row=2, values_only=True):
                v = num(r[ci]) if ci < len(r) else None
                if v:
                    by_month[d[:7]] += v
                    days.add(d)
                if ni is not None and ni < len(r):
                    conv_cnt[d[:7]] += num(r[ni]) or 0
                if vi is not None and vi < len(r):
                    conv_rev[d[:7]] += num(r[vi]) or 0
    wb.close()
    stats = {m: {"전환수": int(conv_cnt[m]), "전환매출": int(conv_rev[m])} for m in by_month}
    return {k: int(round(v)) for k, v in by_month.items()}, sorted(days), stats


def main():
    files = [p for p in SRC.glob("*.xlsx") if not p.name.startswith("~$")]
    if not files:
        sys.exit(f"[오류] 엑셀 파일이 없습니다. 다음 폴더에 넣어 주세요:\n       {SRC}")

    # 매출 기간에서 연도를 가져옵니다 (제목에 연도가 없어서)
    year = date.today().year
    sp = DATA / "sales.csv"
    if sp.exists():
        with open(sp, newline="", encoding="utf-8-sig") as f:
            ds = [r[list(r.keys())[0]] for r in csv.DictReader(f)]
        if ds:
            year = int(min(ds)[:4])

    rows, all_blocks, ad_months, ad_days, ad_stats = [], [], {}, [], {}
    for p in files:
        low = p.name
        if "광고" in low:
            ad_months, ad_days, ad_stats = read_ads(p)
            print(f"광고 리포트 읽음: {p.name} · {len(ad_days)}일 · {min(ad_days)} ~ {max(ad_days)}")
        elif "고정" in low or "마케팅" in low or "비용" in low:
            got, blocks = read_fixed(p, year)
            rows += got
            all_blocks += blocks
            print(f"고정지출 시트 읽음: {p.name}")
        else:
            print(f"[안내] 종류를 알 수 없어 건너뜁니다: {p.name}")
            print(f"       파일 이름에 '고정' 또는 '광고'를 넣어 주세요.")

    if not rows and not ad_months:
        sys.exit("[오류] 읽어 들인 비용이 없습니다. 파일 형식을 확인해 주세요.")

    for m, cost in sorted(ad_months.items()):
        rows.append((m, "마케팅비", "네이버 검색광고", cost, AD_ACCOUNT))

    with open(DATA / "costs.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["월", "구분", "항목", "계정", "금액"])
        for r in sorted(rows, key=lambda x: (x[0], x[1], x[2])):
            month, kind, name, amt = r[0], r[1], r[2], r[3]
            acct = r[4] if len(r) > 4 else ""
            w.writerow([month, kind, name, acct, amt])

    if ad_stats:
        with open(DATA / "광고지표.csv", "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["월", "계정", "광고비", "광고전환수", "광고전환매출"])
            for m in sorted(ad_stats):
                w.writerow([m, AD_ACCOUNT, ad_months[m], ad_stats[m]["전환수"], ad_stats[m]["전환매출"]])

    print()
    for title, kind, mons, items in all_blocks:
        tot = sum(a for _, a in items)
        print(f"[{title}] {kind} · {len(items)}개 항목 · 월 ₩{tot:,} · 적용 {mons[0]} ~ {mons[-1]}")
    if ad_months:
        print(f"\n[네이버 검색광고] 월별 (계정: {AD_ACCOUNT})")
        for m in sorted(ad_months):
            print(f"   {m}  ₩{ad_months[m]:>12,}")
    print(f"\ncosts.csv 저장 완료 · {len(rows)}줄")
    print("주의: 고정지출과 그레비테이트·최과장 마케팅비는 회사 전체 비용입니다.")
    print("      계정 칸을 비워 두어 '공통 비용'으로 처리했습니다.")


if __name__ == "__main__":
    main()
