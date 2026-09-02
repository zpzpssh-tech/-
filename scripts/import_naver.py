"""
네이버 정산내역(SettleCaseByCase) 엑셀 → sales.csv 변환기
--------------------------------------------------------
쓰는 법:
  1) 네이버 판매자센터에서 받은 '건별 정산내역' 엑셀을 data/원본/네이버/ 폴더에 넣습니다.
     (파일 이름은 아무거나 상관없습니다. 2026-06.xlsx 처럼 월별로 두면 관리가 편합니다.)
  2) python3 scripts/import_naver.py
  3) python3 scripts/build.py

매출 계산 규칙 (대표님 지시 기준):
  · 매출 = '정산예정금액'  → 네이버 수수료가 이미 빠진 실수령 기준 금액입니다.
  · '정산전 취소'  → 아예 제외. 돈을 받은 적이 없는 건이라 매출에 잡지 않습니다.
  · '빠른정산 회수' → 그대로 반영(파일에 이미 마이너스). 먼저 받은 돈을 네이버가 도로 가져간 건.
  · '정산후 취소'  → 그대로 반영(파일에 이미 마이너스). 정산 후 환불해 준 건.
  · '빠른정산', '일반정산' → 그대로 플러스 반영.
  · 날짜 기준은 '결제일'입니다.

주의: 정산 파일에는 '판매 수량' 칸이 없습니다. 그래서 상품주문 한 줄을 1개로 셉니다.
      옵션으로 2개를 산 주문은 실제보다 적게 잡힐 수 있습니다.
"""
import csv
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("[오류] openpyxl이 없습니다. 터미널에서 'pip install openpyxl' 을 먼저 실행해 주세요.")

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "원본" / "네이버"
DATA = ROOT / "data"
ACCOUNT = "네이버"
CHANNEL = "네이버"

# 매출에서 제외할 정산상태 (돈을 받은 적이 없는 건)
EXCLUDE_STATES = {"정산전 취소"}
# 알고 있는 정산상태. 새로운 상태가 나오면 경고하고 멈춥니다.
KNOWN_STATES = {"빠른정산", "일반정산", "빠른정산 회수", "정산후 취소", "정산전 취소"}

REQUIRED_COLS = ["주문번호", "상품주문번호", "구분", "상품명", "결제일", "정산상태", "정산예정금액", "정산기준금액"]

# 상품명에서 부위를 찾아내는 규칙. 위에서부터 먼저 맞는 것으로 정합니다.
CATEGORY_RULES = [
    ("무릎",   ["무릎", "무릅", "무름", "슬개", "니슬리브", "연골"]),
    ("허리",   ["허리", "요추", "복대", "척추", "디스크", "갈비뼈"]),
    ("손목",   ["손목", "팔목", "TFCC"]),
    ("팔꿈치", ["팔꿈치", "엘보", "전완근"]),
    ("발/발목", ["발목", "발바닥", "발아치", "아치", "깔창", "뒤꿈치", "발가락", "주상골"]),
    ("어깨/목", ["어깨", "회전근개", "쇄골", "경추", "목 보호대", "라운드숄더", "굽은등", "자세"]),
    ("손가락", ["손가락", "엄지", "검지", "중지", "약지", "새끼"]),
    ("기타용품", ["마사지볼", "마사지 볼", "악력기", "테이프", "테이핑", "파스", "겔", "쇼핑백", "스트랩 교체", "교체 스트랩", "땅콩"]),
]
# 상품명 칸에 옵션값만 들어온 줄(예: "블랙", "L", "1쌍")을 알아내는 규칙
OPTION_ONLY = re.compile(
    r"^(블랙|블루|레드|베이지|화이트|그레이|핑크|네이비|옐로우|하드\s|롤형|튜브형|"
    r"[SMLX]{1,3}(\(|$)|\d+\s?(p|P|쌍|개)|.{0,12}(한쪽|양쪽|한팔|1쌍|1set|1P|2p))"
)


def categorize(name: str) -> str:
    for cat, keys in CATEGORY_RULES:
        if any(k in name for k in keys):
            return cat
    return "미분류"


def short_name(name: str) -> str:
    """긴 검색용 상품명을 화면에 쓸 짧은 이름으로 줄입니다."""
    s = name.replace("물리치료사가 판매하는", "").replace("물리치료사가 직접 판매하는", "")
    s = s.replace("올투게더나우", "").strip()
    s = re.sub(r"^(변경|수정|이름변경)\s*[:：]\s*", "", s).strip()
    s = re.sub(r"\s+", " ", s)
    words = s.split()
    return " ".join(words[:6]) if words else name


def read_sheet(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    ws.reset_dimensions()          # 네이버 파일은 크기 정보가 비어 있어 다시 계산해야 합니다
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows)]
    except StopIteration:
        wb.close()
        sys.exit(f"[오류] 빈 파일입니다: {path.name}")
    missing = [c for c in REQUIRED_COLS if c not in header]
    if missing:
        wb.close()
        sys.exit(f"[오류] {path.name}에 필요한 열이 없습니다: {', '.join(missing)}\n"
                 f"       네이버 판매자센터 > 정산관리 > 건별 정산내역 엑셀이 맞는지 확인해 주세요.")
    idx = {c: header.index(c) for c in header}
    width = len(header)
    for r in rows:
        r = list(r) + [None] * (width - len(r))
        if r[0] is None or str(r[0]).strip() == "":
            continue
        yield idx, r
    wb.close()


def norm_date(v) -> str:
    s = str(v).strip()[:10].replace(".", "-").replace("/", "-")
    parts = s.split("-")
    if len(parts) != 3:
        return ""
    y, m, d = parts
    return f"{y}-{int(m):02d}-{int(d):02d}" if y.isdigit() and m.isdigit() and d.isdigit() else ""


def load_name_to_pid():
    """import_costs.py가 만든 상품매핑.csv를 읽습니다.
    같은 상품이 시기에 따라 이름이 바뀌어도 상품번호 하나로 묶기 위한 표입니다."""
    path = DATA / "상품매핑.csv"
    if not path.exists():
        print("[안내] data/상품매핑.csv 가 없어 상품명 기준으로 묶습니다.")
        print("       scripts/import_costs.py 를 먼저 돌리면 상품번호로 정확히 묶입니다.")
        return {}
    m = {}
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            nm = (r.get("상품명") or "").strip()
            no = (r.get("상품번호") or "").strip()
            if nm and no:
                m[nm] = no
    return m


def load_existing_products():
    """이미 만들어 둔 products.csv의 상품코드와 원가를 그대로 이어받습니다."""
    path = DATA / "products.csv"
    if not path.exists():
        return {}, 0
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    reg, maxno = {}, 0
    for r in rows:
        full = (r.get("정산상품명") or r.get("상품명") or "").strip()
        if not full:
            continue
        r.setdefault("정산상품명", full)
        r.setdefault("상품번호", "")
        r.setdefault("원가", "")
        r.setdefault("포장비", "")
        reg[full] = r
        m = re.match(r"NV-(\d+)$", (r.get("상품코드") or "").strip())
        if m:
            maxno = max(maxno, int(m.group(1)))
    return reg, maxno


def main():
    files = sorted(p for p in SRC.glob("*.xlsx") if not p.name.startswith("~$"))
    if not files:
        sys.exit(f"[오류] 엑셀 파일이 없습니다. 다음 폴더에 정산 엑셀을 넣어 주세요:\n       {SRC}")

    name2pid = load_name_to_pid()
    registry, maxno = load_existing_products()
    pid_code = {}                      # 상품번호 → 이미 붙은 상품코드
    for rec in registry.values():
        if rec.get("상품번호"):
            pid_code[rec["상품번호"]] = rec["상품코드"]
    agg = defaultdict(lambda: {"orders": set(), "qty": 0, "rev": 0, "gross": 0})
    orders_by_date = defaultdict(set)     # 날짜별 주문번호 (상품이 여러 개인 주문을 1건으로)
    buyer_orders = defaultdict(set)       # 구매자명 → 주문번호 (재구매율 계산용)
    cancel_orders = set()                 # 정산전 취소가 있었던 주문번호
    seen = set()
    name_hits = Counter()          # (상품코드, 정산상품명) → 등장 횟수
    stats = defaultdict(lambda: defaultdict(int))   # 파일별 정산상태 집계
    dupes = 0
    unknown_states = set()

    for path in files:
        for idx, r in read_sheet(path):
            state = str(r[idx["정산상태"]] or "").strip()
            if state not in KNOWN_STATES:
                unknown_states.add(state)
                continue
            kind = str(r[idx["구분"]] or "").strip()
            amount = r[idx["정산예정금액"]] or 0
            gross = r[idx["정산기준금액"]] or 0     # 수수료 차감 전 금액
            stats[path.name][state] += 1
            stats[path.name]["_금액_" + state] += amount

            if state in EXCLUDE_STATES:
                cancel_orders.add(str(r[idx["주문번호"]]))
                continue

            date = norm_date(r[idx["결제일"]])
            if not date:
                continue

            # 같은 파일을 두 번 넣었을 때 중복으로 더해지지 않게 막습니다
            key = (str(r[idx["주문번호"]]), str(r[idx["상품주문번호"]]), kind, state,
                   str(r[idx.get("정산기준일", idx["결제일"])]), amount)
            if key in seen:
                dupes += 1
                continue
            seen.add(key)

            raw_name = str(r[idx["상품명"]] or "").strip()
            if kind == "배송비":
                code, full = "SHIP", "배송비"
            else:
                no = name2pid.get(raw_name, "")
                if no:
                    # 상품번호로 묶습니다. 이름이 바뀌어도 같은 상품으로 취급합니다.
                    code = pid_code.get(no)
                    if not code:
                        maxno += 1
                        code = f"NV-{maxno:03d}"
                        pid_code[no] = code
                    full = raw_name
                    rec = registry.get(full)
                    if not rec or rec["상품코드"] != code:
                        registry[full] = {"상품코드": code, "상품명": short_name(raw_name),
                                          "카테고리": categorize(raw_name), "정산상품명": full,
                                          "상품번호": no,
                                          "원가": (rec or {}).get("원가", ""),
                                          "포장비": (rec or {}).get("포장비", "")}
                    else:
                        rec["상품번호"] = no
                elif not raw_name or OPTION_ONLY.match(raw_name):
                    code, full = "OPT", "옵션·추가구성"      # 상품명 칸에 옵션값만 들어온 줄
                else:
                    full = raw_name
                    if full in registry:
                        code = registry[full]["상품코드"]
                    else:
                        maxno += 1
                        code = f"NV-{maxno:03d}"
                        registry[full] = {"상품코드": code, "상품명": short_name(full),
                                          "카테고리": categorize(full), "정산상품명": full,
                                          "상품번호": "", "원가": "", "포장비": ""}

            orders_by_date[date].add(str(r[idx["주문번호"]]))
            buyer = str(r[idx["구매자명"]] or "").strip()
            if buyer:
                buyer_orders[buyer].add(str(r[idx["주문번호"]]))
            name_hits[(code, full)] += 1
            a = agg[(date, code)]
            a["rev"] += amount
            a["gross"] += gross
            a["orders"].add(str(r[idx["주문번호"]]))
            if kind != "배송비":
                a["qty"] += 1 if amount >= 0 else -1

    if unknown_states:
        sys.exit(f"[오류] 처음 보는 정산상태가 있습니다: {', '.join(sorted(unknown_states))}\n"
                 f"       이 건을 매출에 넣을지 뺄지 알려주시면 규칙에 추가하겠습니다.")

    # 고정 항목(배송비·옵션) 등록
    for code, name, cat in [("SHIP", "배송비", "배송비"), ("OPT", "옵션·추가구성", "미분류")]:
        full = name
        if full not in registry:
            registry[full] = {"상품코드": code, "상품명": name, "카테고리": cat,
                              "정산상품명": full, "상품번호": "", "원가": "", "포장비": ""}

    # ── sales.csv 저장 ──
    out = []
    for (date, code), a in sorted(agg.items()):
        out.append([date, CHANNEL, ACCOUNT, code, len(a["orders"]), a["qty"], round(a["rev"]), round(a["gross"])])
    with open(DATA / "sales.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["날짜", "채널", "계정", "상품코드", "주문건수", "판매수량", "매출", "총매출"])
        w.writerows(out)

    with open(DATA / "일별주문.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["날짜", "계정", "주문건수"])
        for d in sorted(orders_by_date):
            w.writerow([d, ACCOUNT, len(orders_by_date[d])])

    # ── products.csv 저장 ──
    # 원가·포장비는 대표님이 채우는 칸입니다. 맨 뒤 '참고_' 칸은 자동으로 다시 채워지니
    # 고쳐도 소용없고, 어느 상품부터 원가를 넣어야 할지 고르는 용도로만 보시면 됩니다.
    # 매출이 큰 상품이 위로 오게 정렬합니다.
    totals = defaultdict(lambda: {"rev": 0, "qty": 0})
    for (_, code), a in agg.items():
        totals[code]["rev"] += a["rev"]
        totals[code]["qty"] += a["qty"]
    # 상품코드 하나에 이름이 여러 개일 수 있으므로 대표 한 줄만 남깁니다.
    by_code = {}
    aliases = defaultdict(list)
    for rec in registry.values():
        code = rec["상품코드"]
        aliases[code].append(rec["정산상품명"])
        cur = by_code.get(code)
        # 가장 자주 쓰인 이름을 대표로 씁니다 ("변경 : ..." 같은 임시 이름이 뽑히지 않게)
        if cur is None or name_hits[(code, rec["정산상품명"])] > name_hits[(code, cur["정산상품명"])]:
            by_code[code] = rec
    prods = sorted(by_code.values(), key=lambda p: -totals[p["상품코드"]]["rev"])
    with open(DATA / "products.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["상품코드", "상품명", "카테고리", "상품번호", "원가", "포장비",
                    "정산상품명", "참고_누적매출", "참고_판매수량", "참고_이름수"])
        for p in prods:
            t = totals[p["상품코드"]]
            w.writerow([p["상품코드"], p["상품명"], p["카테고리"], p.get("상품번호", ""),
                        p.get("원가", ""), p.get("포장비", ""), p["정산상품명"],
                        round(t["rev"]), t["qty"], len(set(aliases[p["상품코드"]]))])

    # ── 고객지표.csv 저장 (구매자명 기준. 동명이인은 한 사람으로 세므로 대략치입니다) ──
    total_orders = sum(len(v) for v in orders_by_date.values())
    buyers = len(buyer_orders)
    repeat_buyers = sum(1 for v in buyer_orders.values() if len(v) >= 2)
    repeat_orders = sum(len(v) - 1 for v in buyer_orders.values() if len(v) >= 2)
    with open(DATA / "고객지표.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["항목", "값"])
        w.writerow(["구매자수", buyers])
        w.writerow(["재구매자수", repeat_buyers])
        w.writerow(["재구매주문수", repeat_orders])
        w.writerow(["총주문수", total_orders])
        w.writerow(["취소주문수", len(cancel_orders)])
        w.writerow(["1인당평균주문", round(total_orders / buyers, 3) if buyers else 0])

    # ── 정산요약.csv 저장 (규칙이 제대로 적용됐는지 확인용) ──
    summary = []
    for name in sorted(stats):
        s = stats[name]
        row = {"파일": name}
        for st in ["빠른정산", "일반정산", "빠른정산 회수", "정산후 취소", "정산전 취소"]:
            row[st + "_건수"] = s.get(st, 0)
            row[st + "_금액"] = round(s.get("_금액_" + st, 0))
        row["매출반영"] = sum(row[st + "_금액"] for st in ["빠른정산", "일반정산", "빠른정산 회수", "정산후 취소"])
        summary.append(row)
    with open(DATA / "정산요약.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(summary[0].keys()))
        w.writeheader()
        w.writerows(summary)

    # ── 결과 보고 ──
    total = sum(r[6] for r in out)
    gross_total = sum(r[7] for r in out)
    dates = sorted({r[0] for r in out})
    print(f"파일 {len(files)}개 읽음: {', '.join(p.name for p in files)}")
    print(f"기간 {dates[0]} ~ {dates[-1]} · {len(dates)}일 · 상품 {len(prods)}개")
    if dupes:
        print(f"중복으로 걸러낸 줄: {dupes}개")
    print("\n[월별 정산상태 확인]")
    print(f"{'파일':<14}{'매출로 잡은 금액':>18}{'제외(정산전취소)':>18}{'회수·취소 마이너스':>20}")
    for name in sorted(stats):
        s = stats[name]
        included = sum(v for k, v in s.items() if k.startswith("_금액_") and k[4:] not in EXCLUDE_STATES)
        excluded = sum(v for k, v in s.items() if k.startswith("_금액_") and k[4:] in EXCLUDE_STATES)
        minus = s.get("_금액_빠른정산 회수", 0) + s.get("_금액_정산후 취소", 0)
        print(f"{name:<14}{included:>18,}{excluded:>18,}{minus:>20,}")
    print(f"\n총 매출(정산예정금 기준): ₩{total:,}")
    print(f"총매출(수수료 차감 전): ₩{gross_total:,} · 네이버 수수료 ₩{gross_total - total:,} "
          f"({(gross_total - total) / gross_total * 100:.1f}%)")
    print(f"구매자 {buyers:,}명 · 재구매 고객 {repeat_buyers:,}명 "
          f"({repeat_buyers / buyers * 100:.1f}%) · 1인당 평균 {total_orders / buyers:.2f}회")
    if not (DATA / "원가일계.csv").exists():
        print("\n※ 원가 데이터가 없습니다. scripts/import_costs.py 를 먼저 실행하면 마진까지 계산됩니다.")


if __name__ == "__main__":
    main()
