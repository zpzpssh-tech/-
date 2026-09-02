"""
쿠팡 로켓배송(직매입) 발주서를 읽어 하루치 매출을 만듭니다.
------------------------------------------------------------
넣는 곳 : data/원본/로켓배송/*.xlsx  (쿠팡 WING에서 받은 발주서 파일 그대로)

로켓배송은 쿠팡이 우리 물건을 사가는 방식(직매입)이라 판매수수료가 없습니다.
발주서의 '입고금액 > 매입가'가 곧 우리 매출이고, 부가세가 포함된 금액입니다.

  발주금액 = 쿠팡이 주문한 금액
  입고금액 = 우리가 실제로 납품한 금액   ← 이것만 매출로 잡습니다
  둘의 차이 = 재고가 없어 못 보낸 금액 (기회손실)

날짜는 실제 납품일(하차일시)을 씁니다. 없으면 입고예정일을 씁니다.

원가는 쿠팡 계정 엑셀들의 '원가' 시트를 모아 상품명으로 붙입니다.
못 찾은 것은 data/로켓_원가미매칭.csv 에 적어 둡니다.

만드는 파일
  data/로켓_일별.csv       날짜, 계정, 발주건수, 판매량, 총매출, 원가, 원가미상수량, 발주금액, 미납금액
  data/로켓_상품.csv       월, 상품코드, 상품명, 발주수량, 입고수량, 총매출, 원가, 미납금액
  data/로켓_원가미매칭.csv  상품명, 수량, 매출
"""
import csv
import importlib.util
import re
import warnings
from collections import defaultdict
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore", module="openpyxl")

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "원본" / "로켓배송"
OUT = ROOT / "data"
ACCOUNT = "로켓배송"


def num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d.\-]", "", str(v))
    return float(s) if s not in ("", "-", ".") else 0.0


def as_date(v):
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    m = re.match(r"(\d{4})[/-](\d{2})[/-](\d{2})", str(v or ""))
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else ""


def load_costs():
    """쿠팡 계정 엑셀들의 '원가' 시트를 모아 (옵션명→원가, 상품명→{꼬리표:원가}) 를 만듭니다."""
    spec = importlib.util.spec_from_file_location("cp", ROOT / "scripts" / "import_coupang.py")
    cp = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(cp)
    by_opt, by_prod = {}, defaultdict(dict)
    for path in sorted((ROOT / "data" / "원본" / "쿠팡").glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "원가" not in wb.sheetnames:
            continue
        ws = wb["원가"]
        ws.reset_dimensions()
        it = ws.iter_rows(values_only=True)
        next(it, None)
        for r in it:
            r = list(r) + [None] * 3
            if r[0] is None:
                continue
            opt, prd, c = str(r[0]).strip(), str(r[1] or "").strip(), num(r[2])
            if not c:                       # 0원은 '아직 안 채운 칸'이라 건너뜁니다
                continue
            by_opt.setdefault(opt, c)
            if prd:
                by_prod[prd].setdefault(cp.tail(opt), c)
    # 대표가 직접 적어준 원가는 시트보다 우선합니다 (data/로켓_원가.csv: 상품명, 원가, 메모)
    add = ROOT / "data" / "로켓_원가.csv"
    if add.exists():
        with open(add, newline="", encoding="utf-8-sig") as f:
            for r in csv.DictReader(f):
                k = {(c or "").lstrip("\ufeff").strip(): v for c, v in r.items()}
                nm, c = (k.get("상품명") or "").strip(), num(k.get("원가"))
                if nm and c:
                    by_opt[nm] = c

    def find(name):
        """상품명 전체 → (원가, 어떻게 찾았는지). 못 찾으면 (None, '못 찾음')"""
        if name in by_opt:
            return by_opt[name], "옵션명 일치"
        prd = name.split(",")[0].strip()
        d = by_prod.get(prd)
        if not d:
            return None, "못 찾음"
        t = cp.tail(name)
        want = cp.opt_qty(t)
        if t in d:
            return d[t], "상품명+꼬리표"
        if len(d) == 1:
            key, val = next(iter(d.items()))
            have = cp.opt_qty(key)
            if want != have and have:
                return val * want / have, "수량 배수"
            return val, "상품명 원가 1개"
        same = [v for k, v in d.items() if cp.opt_qty(k) == want]
        if same:
            return sum(same) / len(same), "수량 같은 항목"
        unit = sum(v / max(1, cp.opt_qty(k)) for k, v in d.items()) / len(d)
        return unit * want, "평균 단가 × 수량"

    return find


def read_po(path):
    """발주서 한 장 → (발주번호, 납품일, [품목...])"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    ws.reset_dimensions()
    rows = [list(r) + [None] * 22 for r in ws.iter_rows(values_only=True)]

    po, date, head, foot = "", "", None, None
    for i, r in enumerate(rows):
        c0 = str(r[0] or "").strip()
        if c0.startswith("발주서 No."):
            po = c0.replace("발주서 No.", "").strip()
        # 날짜 머리글 다음 줄에 [.., 물류센터, 주소, 입고예정일, 하차일시] 가 옵니다
        if c0 == "입고예정일시" and str(r[6] or "").strip() == "하차일시":
            nxt = rows[i + 1]
            date = as_date(nxt[6]) or as_date(nxt[5])
        if c0 == "No." and str(r[1] or "").strip() == "상품코드":
            head = i
        if c0 == "합계" and head is not None and foot is None:
            foot = i
    if head is None or foot is None:
        return po, date, None

    items = []
    for r in rows[head + 2:foot]:
        if not str(r[0] or "").strip().isdigit():
            continue
        items.append({
            "코드": str(r[1] or "").strip(),
            "상품명": str(r[2] or "").strip(),
            "발주수량": num(r[6]), "입고수량": num(r[8]),
            "단가": num(r[9]), "발주금액": num(r[12]), "매출": num(r[16]),
        })
    return po, date, items


def main():
    if not SRC.exists():
        print(f"로켓배송 발주서 폴더가 없습니다 → {SRC}")
        return

    find_cost = load_costs()
    daily = defaultdict(lambda: {"발주건수": 0, "판매량": 0, "총매출": 0.0, "원가": 0.0,
                                 "원가미상수량": 0, "발주금액": 0.0, "미납금액": 0.0})
    prod = defaultdict(lambda: {"발주수량": 0, "입고수량": 0, "총매출": 0.0,
                                "원가": 0.0, "미납금액": 0.0})
    unmatched = defaultdict(lambda: [0, 0.0])
    names = {}
    seen, skipped, nodate = set(), [], 0

    for path in sorted(SRC.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        po, date, items = read_po(path)
        if items is None:
            skipped.append(path.name)
            continue
        if po in seen:          # 같은 발주서를 두 번 받아도 한 번만 셉니다
            continue
        seen.add(po)
        if not date:
            nodate += 1
            continue
        d = daily[date]
        d["발주건수"] += 1
        for it in items:
            gap = it["발주금액"] - it["매출"]
            unit, _ = find_cost(it["상품명"])
            cogs = (unit or 0) * it["입고수량"]
            d["판매량"] += it["입고수량"]
            d["총매출"] += it["매출"]
            d["원가"] += cogs
            d["발주금액"] += it["발주금액"]
            d["미납금액"] += gap
            if unit is None and it["입고수량"]:
                d["원가미상수량"] += it["입고수량"]
                u = unmatched[it["상품명"]]
                u[0] += it["입고수량"]
                u[1] += it["매출"]
            key = (date[:7], it["코드"])
            p = prod[key]
            p["발주수량"] += it["발주수량"]
            p["입고수량"] += it["입고수량"]
            p["총매출"] += it["매출"]
            p["원가"] += cogs
            p["미납금액"] += gap
            names[it["코드"]] = it["상품명"]

    with open(OUT / "로켓_일별.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["날짜", "계정", "발주건수", "판매량", "총매출", "원가", "원가미상수량",
                    "발주금액", "미납금액"])
        for d in sorted(daily):
            v = daily[d]
            w.writerow([d, ACCOUNT, v["발주건수"], round(v["판매량"]), round(v["총매출"]),
                        round(v["원가"]), round(v["원가미상수량"]),
                        round(v["발주금액"]), round(v["미납금액"])])

    with open(OUT / "로켓_상품.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["월", "상품코드", "상품명", "발주수량", "입고수량", "총매출", "원가", "미납금액"])
        for (m, code) in sorted(prod):
            p = prod[(m, code)]
            w.writerow([m, code, names[code], round(p["발주수량"]), round(p["입고수량"]),
                        round(p["총매출"]), round(p["원가"]), round(p["미납금액"])])

    with open(OUT / "로켓_원가미매칭.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["상품명", "수량", "매출"])
        for nm, v in sorted(unmatched.items(), key=lambda x: -x[1][1]):
            w.writerow([nm, round(v[0]), round(v[1])])

    print(f"로켓배송 발주서 {len(seen)}건 → data/로켓_일별.csv")
    if skipped:
        print(f"  [건너뜀] 형식이 다른 파일 {len(skipped)}개: {', '.join(skipped[:3])}")
    if nodate:
        print(f"  [건너뜀] 납품일이 없는 발주서 {nodate}건")

    by_m = defaultdict(lambda: [0.0, 0.0, 0.0, 0])
    for d, v in daily.items():
        t = by_m[d[:7]]
        t[0] += v["발주금액"]; t[1] += v["총매출"]; t[2] += v["미납금액"]; t[3] += v["판매량"]
    print(f"  {'월':<9}{'발주금액':>13}{'매출(입고)':>13}{'미납':>12}{'납품률':>8}{'수량':>8}")
    T = [0.0, 0.0, 0.0, 0]
    for m in sorted(by_m):
        v = by_m[m]
        for i in range(4):
            T[i] += v[i]
        print(f"  {m:<9}{v[0]:>13,.0f}{v[1]:>13,.0f}{v[2]:>12,.0f}"
              f"{v[1] / v[0] * 100 if v[0] else 0:>7.0f}%{v[3]:>8,.0f}")
    print(f"  {'합계':<9}{T[0]:>13,.0f}{T[1]:>13,.0f}{T[2]:>12,.0f}"
          f"{T[1] / T[0] * 100 if T[0] else 0:>7.0f}%{T[3]:>8,.0f}  (부가세 포함)")

    qty = sum(v["판매량"] for v in daily.values())
    unk = sum(v["원가미상수량"] for v in daily.values())
    cogs = sum(v["원가"] for v in daily.values())
    print(f"  원가 {cogs:,.0f} · 반영률 {(qty - unk) / qty * 100 if qty else 0:.1f}% "
          f"(원가 못 찾은 수량 {unk:,.0f}개)")
    if unmatched:
        print(f"  원가를 못 찾은 상품 {len(unmatched)}개 → data/로켓_원가미매칭.csv")
        for nm, v in sorted(unmatched.items(), key=lambda x: -x[1][1])[:6]:
            print(f"    {v[0]:>6,.0f}개  매출 {v[1]:>11,.0f}  {nm[:46]}")


if __name__ == "__main__":
    main()
