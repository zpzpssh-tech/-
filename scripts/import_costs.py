"""
원가관리 엑셀 → 원가·수량 데이터 변환기
---------------------------------------
쓰는 법:
  1) 원가관리 엑셀을 data/원본/원가/ 폴더에 넣습니다.
  2) python3 scripts/import_costs.py
  3) python3 scripts/import_naver.py
  4) python3 scripts/build.py

엑셀 구조 (이 형식을 그대로 씁니다):
  [원가] 시트      상품번호 | 상품명 | 옵션정보 | 원가
  [YYYYMM월] 시트  주문일시 | 주문상태 | 상품번호 | 상품명 | 옵션정보 | 수량

주문 상세 파일 (선택):
  data/원본/주문/*.xlsx 에 네이버에서 받은 주문 상세를 넣으면, 그 파일이 덮는 달은
  원가관리 엑셀의 월별 시트 대신 이 파일을 씁니다 (더 최신이고 칸이 많습니다).
  '배송속성'(N배송 / N판매자배송 / 일반배송) 칸이 있으면 날짜별 택배 종류도 뽑아
  data/배송속성.csv 로 저장합니다. 택배비를 실제 단가로 계산하는 데 씁니다.

원가 붙이는 순서 (위에서부터 먼저 맞는 것을 씁니다):
  1. 상품번호 + 옵션정보가 정확히 같을 때
  2. 그 상품번호에 옵션 없는 원가가 있을 때
  3. 색상 표현만 다를 때 (블랙/블루 등을 지우고 비교)
  4. 그 상품번호의 원가가 딱 하나뿐일 때
  못 찾은 건은 data/원가_미매칭.csv 에 수량이 많은 순으로 적어 둡니다.

주문상태 처리:
  · 취소, 반품  → 원가에서 제외 (물건이 돌아온 건)
  · 나머지(구매확정·배송완료·배송중·결제완료·교환) → 원가 발생
"""
import csv
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("[오류] openpyxl이 없습니다. 'pip install openpyxl' 을 먼저 실행해 주세요.")

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "원본" / "원가"
ORDER_SRC = ROOT / "data" / "원본" / "주문"
DATA = ROOT / "data"

EXCLUDE_STATES = {"취소", "반품"}          # 물건이 돌아와 원가가 발생하지 않는 상태
COLOR = re.compile(r"(블랙|블루|레드|그레이|화이트|베이지|핑크|네이비|스킨|회색|검정|옐로우|퍼플|민트)")


def pid(v) -> str:
    """상품번호를 문자열로 통일합니다 (엑셀에서 5799070116.0 처럼 읽히는 것 정리)."""
    return str(v).split(".")[0].strip() if v is not None else ""


def norm_opt(o: str) -> str:
    """색상 표현을 지우고 수량·사이즈만 남겨 비교용으로 만듭니다."""
    return COLOR.sub("", (o or "").replace(" ", ""))


def main():
    files = sorted(p for p in SRC.glob("*.xlsx") if not p.name.startswith("~$"))
    if not files:
        sys.exit(f"[오류] 엑셀 파일이 없습니다. 원가관리 엑셀을 여기에 넣어 주세요:\n       {SRC}")
    if len(files) > 1:
        print(f"[안내] 파일이 여러 개라 가장 최근 것을 씁니다: {files[-1].name}")
    path = files[-1]

    wb = openpyxl.load_workbook(path, data_only=True)
    if "원가" not in wb.sheetnames:
        sys.exit(f"[오류] '원가' 시트가 없습니다. 현재 시트: {', '.join(wb.sheetnames)}")

    # ── 원가 시트 ──
    cost = {}                       # (상품번호, 옵션) → 원가
    cost_by_no = defaultdict(dict)
    for i, r in enumerate(wb["원가"].iter_rows(min_row=2, values_only=True), start=2):
        if r[0] is None:
            continue
        no, opt = pid(r[0]), (str(r[2]).strip() if r[2] else "")
        if r[3] is None or str(r[3]).strip() == "":
            continue
        try:
            v = int(round(float(str(r[3]).replace(",", ""))))
        except ValueError:
            sys.exit(f"[오류] 원가 시트 {i}행의 원가가 숫자가 아닙니다: '{r[3]}'")
        cost[(no, opt)] = v
        cost_by_no[no][opt] = v
    if not cost:
        sys.exit("[오류] '원가' 시트에서 원가를 한 건도 읽지 못했습니다.")

    # ── data/원가_추가.csv: 원가관리 엑셀에 아직 없는 옵션을 손으로 채워 넣는 파일 ──
    extra_path = DATA / "원가_추가.csv"
    extra_n = 0
    if extra_path.exists():
        with open(extra_path, newline="", encoding="utf-8-sig") as f:
            for i, r in enumerate(csv.DictReader(f), start=2):
                keys = {(k or "").lstrip("\ufeff").strip(): (v if isinstance(v, str) else "")
                        for k, v in r.items()}
                no = pid(keys.get("상품번호"))
                opt = keys.get("옵션정보", "").strip()
                raw = keys.get("원가", "").replace(",", "").strip()
                if not no or raw == "":
                    continue
                try:
                    v = int(round(float(raw)))
                except ValueError:
                    sys.exit(f"[오류] 원가_추가.csv {i}행의 원가가 숫자가 아닙니다: '{raw}'")
                cost[(no, opt)] = v
                cost_by_no[no][opt] = v
                extra_n += 1

    norm_by_no = defaultdict(dict)
    for no, opts in cost_by_no.items():
        for o, v in opts.items():
            norm_by_no[no].setdefault(norm_opt(o), v)

    def find_cost(no, opt):
        if (no, opt) in cost:
            return cost[(no, opt)], "정확히 일치"
        if (no, "") in cost:
            return cost[(no, "")], "옵션없는 원가"
        n = norm_by_no.get(no, {})
        if norm_opt(opt) in n:
            return n[norm_opt(opt)], "색상 무시 일치"
        if len(cost_by_no.get(no, {})) == 1:
            return next(iter(cost_by_no[no].values())), "그 번호 원가 1개"
        return None, "못 찾음"

    # ── 주문 상세 파일 (있으면 그 달은 이쪽을 씁니다) ──
    NEED = ["주문일시", "주문상태", "상품번호", "상품명", "옵션정보", "수량"]
    order_rows, ship_rows, covered = [], [], set()
    if ORDER_SRC.exists():
        for path2 in sorted(ORDER_SRC.glob("*.xlsx")):
            if path2.name.startswith("~$"):
                continue
            wb2 = openpyxl.load_workbook(path2, read_only=True, data_only=True)
            for sh2 in wb2.sheetnames:
                ws2 = wb2[sh2]
                ws2.reset_dimensions()
                it2 = ws2.iter_rows(values_only=True)
                head2 = [str(h).strip() if h else "" for h in (next(it2, None) or [])]
                if any(c not in head2 for c in NEED):
                    continue
                J = {c: head2.index(c) for c in NEED}
                for c in ("배송속성", "주문번호"):
                    if c in head2:
                        J[c] = head2.index(c)
                for r in it2:
                    if r is None or r[J["주문일시"]] is None:
                        continue
                    d = str(r[J["주문일시"]])[:10].replace(".", "-").replace("/", "-")
                    order_rows.append((d, [r[J[c]] for c in NEED]))
                    covered.add(d[:7])
                    if "배송속성" in J and "주문번호" in J:
                        ship_rows.append((d, str(r[J["주문번호"]]).strip(),
                                          str(r[J["주문상태"]] or "").strip(),
                                          str(r[J["배송속성"]] or "").strip()))
            wb2.close()
        if covered:
            print(f"주문 상세 파일 {len(order_rows):,}줄 · 덮는 달 {', '.join(sorted(covered))} "
                  f"→ 이 달은 원가관리 엑셀 대신 이 파일을 씁니다")

    # ── 배송속성 (택배 종류) ──
    if ship_rows:
        by_order = {}
        for d, o, state, attr in ship_rows:
            if state == "취소":       # 취소 건은 발송하지 않았습니다
                continue
            by_order[o] = (d, attr)
        cnt = defaultdict(lambda: {"N배송": 0, "판매자배송": 0})
        for d, attr in by_order.values():
            cnt[d]["N배송" if attr == "N배송" else "판매자배송"] += 1
        with open(DATA / "배송속성.csv", "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["날짜", "N배송", "판매자배송"])
            for d in sorted(cnt):
                w.writerow([d, cnt[d]["N배송"], cnt[d]["판매자배송"]])
        tn = sum(v["N배송"] for v in cnt.values())
        ts = sum(v["판매자배송"] for v in cnt.values())
        print(f"배송속성 {len(cnt)}일 · N배송 {tn:,}건 / 판매자배송 {ts:,}건 "
              f"(N배송 {tn / (tn + ts) * 100:.1f}%) → data/배송속성.csv")

    # ── 월별 주문 시트 ──
    sheets = [s for s in wb.sheetnames if re.fullmatch(r"\d{6}월", s)]
    if not sheets:
        sys.exit(f"[오류] 'YYYYMM월' 형태의 월별 시트가 없습니다. 현재 시트: {', '.join(wb.sheetnames)}")

    daily = defaultdict(lambda: {"qty": 0, "cogs": 0, "unknown": 0})   # (날짜, 상품번호)
    opt_month = defaultdict(lambda: {"qty": 0, "cogs": 0})             # (월, 상품번호, 옵션) 재고 예측용
    opt_name = {}
    names = defaultdict(Counter)          # 상품번호 → 상품명 빈도
    tiers = Counter()
    missing = Counter()
    excluded_qty = 0

    for sh in sheets:
        ws = wb[sh]
        head = [str(h).strip() if h else "" for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        need = ["주문일시", "주문상태", "상품번호", "상품명", "옵션정보", "수량"]
        miss = [c for c in need if c not in head]
        if miss:
            sys.exit(f"[오류] '{sh}' 시트에 열이 없습니다: {', '.join(miss)}\n       현재 열: {', '.join(head)}")
        I = {c: head.index(c) for c in need}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[I["주문일시"]] is None:
                continue
            if str(r[I["주문일시"]])[:10].replace(".", "-").replace("/", "-")[:7] in covered:
                continue          # 주문 상세 파일이 있는 달은 건너뜁니다
            state = str(r[I["주문상태"]] or "").strip()
            try:
                qty = int(float(r[I["수량"]] or 0))
            except (TypeError, ValueError):
                continue
            if state in EXCLUDE_STATES:
                excluded_qty += qty
                continue
            date = str(r[I["주문일시"]])[:10].replace(".", "-").replace("/", "-")
            no = pid(r[I["상품번호"]])
            opt = str(r[I["옵션정보"]] or "").strip()
            nm = str(r[I["상품명"]] or "").strip()
            if nm:
                names[no][nm] += 1

            unit, tier = find_cost(no, opt)
            tiers[tier] += qty
            om = opt_month[(date[:7], no, opt)]
            om["qty"] += qty
            om["cogs"] += (unit or 0) * qty
            opt_name[(no, opt)] = nm or opt_name.get((no, opt), "")
            d = daily[(date, no)]
            d["qty"] += qty
            if unit is None:
                d["unknown"] += qty
                missing[(no, opt, nm)] += qty
            else:
                d["cogs"] += unit * qty

    # 주문 상세 파일 줄도 같은 방식으로 원가를 붙입니다
    for date, vals in order_rows:
        state = str(vals[1] or "").strip()
        try:
            qty = int(float(vals[5] or 0))
        except (TypeError, ValueError):
            continue
        if state in EXCLUDE_STATES:
            excluded_qty += qty
            continue
        no = pid(vals[2])
        opt = str(vals[4] or "").strip()
        nm = str(vals[3] or "").strip()
        if nm:
            names[no][nm] += 1
        unit, tier = find_cost(no, opt)
        tiers[tier] += qty
        om = opt_month[(date[:7], no, opt)]
        om["qty"] += qty
        om["cogs"] += (unit or 0) * qty
        opt_name[(no, opt)] = nm or opt_name.get((no, opt), "")
        d = daily[(date, no)]
        d["qty"] += qty
        if unit is None:
            d["unknown"] += qty
            missing[(no, opt, nm)] += qty
        else:
            d["cogs"] += unit * qty
    wb.close()

    # ── 저장 ──
    with open(DATA / "원가일계.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["날짜", "상품번호", "수량", "원가합계", "원가미상수량"])
        for (date, no), d in sorted(daily.items()):
            w.writerow([date, no, d["qty"], d["cogs"], d["unknown"]])

    with open(DATA / "옵션판매.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["월", "상품번호", "옵션정보", "수량", "원가합계"])
        for (m, no, opt), v in sorted(opt_month.items()):
            w.writerow([m, no, opt, v["qty"], v["cogs"]])

    with open(DATA / "상품매핑.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["상품명", "상품번호"])
        seen = set()
        for no, c in names.items():
            for nm in c:
                if (nm, no) not in seen:
                    seen.add((nm, no))
                    w.writerow([nm, no])

    with open(DATA / "원가_미매칭.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["수량", "상품번호", "옵션정보", "상품명", "원가"])
        for (no, opt, nm), q in missing.most_common():
            w.writerow([q, no, opt, nm, ""])

    total_qty = sum(tiers.values())
    known = total_qty - tiers["못 찾음"]
    print(f"원가관리 엑셀 읽음: {path.name} · 월별 시트 {len(sheets)}개")
    print(f"원가 항목 {len(cost)}개 · 상품번호 {len(cost_by_no)}개"
          + (f" (원가_추가.csv에서 {extra_n}개 보탬)" if extra_n else ""))
    print(f"주문 수량 {total_qty:,}개 (취소·반품 {excluded_qty:,}개 제외)\n")
    print("[원가를 어떻게 붙였는지]")
    for k in ["정확히 일치", "옵션없는 원가", "색상 무시 일치", "그 번호 원가 1개", "못 찾음"]:
        if tiers[k]:
            print(f"  {k:<14}{tiers[k]:>9,}개  {tiers[k] / total_qty * 100:>5.1f}%")
    print(f"옵션 조합 {len(opt_month)//max(1,len({k[0] for k in opt_month}))}종 내외 · 옵션판매.csv {len(opt_month)}줄")
    print(f"\n원가 반영률: {known / total_qty * 100:.1f}%  (원가 합계 ₩{sum(d['cogs'] for d in daily.values()):,})")
    if missing:
        print(f"\n※ 원가를 못 찾은 옵션 {len(missing)}종 / {tiers['못 찾음']:,}개")
        print(f"   data/원가_미매칭.csv 에 수량 많은 순으로 적어 뒀습니다.")
        print(f"   원가관리 엑셀의 '원가' 시트에 아래 줄을 추가하시면 100%가 됩니다:")
        for (no, opt, nm), q in missing.most_common(5):
            print(f"     {q:>5}개  상품번호 {no}  옵션[{opt[:44]}]")


if __name__ == "__main__":
    main()
