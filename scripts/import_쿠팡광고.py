"""
쿠팡 광고 리포트(캠페인별 일간 리포트)를 읽어 하루치 광고비를 만듭니다.
------------------------------------------------------------------
넣는 곳 : data/원본/쿠팡광고/<계정이름>/*.xlsx
          예) data/원본/쿠팡광고/올투게더나우/A00254410_pa_daily_campaign_20260101_20260131.xlsx
          폴더 이름이 그대로 계정 이름이 됩니다.

리포트의 '광고비'는 부가세가 빠진 공급가입니다. 우리 장부는 모든 금액을 부가세
포함으로 잡으므로 여기서 1.1을 곱해 넣습니다.

'판매방식' 칸이 3P면 셀러 광고, Retail이면 로켓(retail) 광고입니다.
로켓 매출은 아직 장부에 없어서 따로 표시해 둡니다.

만드는 파일
  data/쿠팡광고.csv      날짜, 계정, 구분, 광고비          ← 하루치 (부가세 포함)
  data/쿠팡광고_캠페인.csv 월, 계정, 구분, 캠페인, 광고비, 노출수, 클릭수, 전환매출, 판매수량
"""
import csv
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "원본" / "쿠팡광고"
OUT = ROOT / "data"
VAT = 1.1  # 리포트는 공급가라 부가세를 붙입니다


def num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d.\-]", "", str(v))
    return float(s) if s not in ("", "-", ".") else 0.0


def as_date(v):
    """20260301 / 2026-03-01 / datetime → '2026-03-01'"""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = re.sub(r"[^\d]", "", str(v or ""))
    return f"{s[:4]}-{s[4:6]}-{s[6:8]}" if len(s) >= 8 else ""


def kind_of(sales_type):
    return "로켓" if str(sales_type or "").strip().lower() == "retail" else "셀러"


def main():
    if not SRC.exists():
        print(f"광고 리포트 폴더가 없습니다 → {SRC}")
        return

    daily = defaultdict(float)     # (날짜, 계정, 구분) → 광고비
    camp = defaultdict(lambda: defaultdict(float))
    months = defaultdict(set)      # 계정 → {월}
    files = 0

    for folder in sorted(p for p in SRC.iterdir() if p.is_dir()):
        account = folder.name
        for path in sorted(folder.glob("*.xlsx")):
            if path.name.startswith("~$"):
                continue
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            ws.reset_dimensions()
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            ix = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
            need = ["날짜", "광고비", "판매방식", "캠페인명"]
            missing = [c for c in need if c not in ix]
            if missing:
                print(f"  [건너뜀] {path.name} — 없는 칸: {', '.join(missing)}")
                continue
            for r in rows:
                if r is None or r[ix["날짜"]] is None:
                    continue
                date = as_date(r[ix["날짜"]])
                if not date:
                    continue
                k = kind_of(r[ix["판매방식"]])
                cost = num(r[ix["광고비"]]) * VAT
                daily[(date, account, k)] += cost
                months[account].add(date[:7])
                c = camp[(date[:7], account, k, str(r[ix["캠페인명"]] or "").strip())]
                c["광고비"] += cost
                for src, dst in (("노출수", "노출수"), ("클릭수", "클릭수"),
                                 ("총 전환매출액(14일)", "전환매출"),
                                 ("총 판매수량(14일)", "판매수량")):
                    if src in ix:
                        c[dst] += num(r[ix[src]])
            files += 1

    with open(OUT / "쿠팡광고.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["날짜", "계정", "구분", "광고비"])
        for (d, a, k), v in sorted(daily.items()):
            w.writerow([d, a, k, round(v)])

    with open(OUT / "쿠팡광고_캠페인.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["월", "계정", "구분", "캠페인", "광고비", "노출수", "클릭수", "전환매출", "판매수량"])
        for (m, a, k, name), c in sorted(camp.items()):
            w.writerow([m, a, k, name.replace(",", " "), round(c["광고비"]),
                        round(c["노출수"]), round(c["클릭수"]),
                        round(c["전환매출"]), round(c["판매수량"])])

    print(f"쿠팡 광고 리포트 {files}개 → data/쿠팡광고.csv")
    for account in sorted(months):
        ms = sorted(months[account])
        print(f"  {account}: {ms[0]} ~ {ms[-1]} ({len(ms)}개월)")
        by_m = defaultdict(lambda: defaultdict(float))
        for (d, a, k), v in daily.items():
            if a == account:
                by_m[d[:7]][k] += v
        print(f"      {'월':<9}{'셀러':>13}{'로켓':>12}{'합계':>13}")
        t_s = t_r = 0.0
        for m in ms:
            s, r = by_m[m]["셀러"], by_m[m]["로켓"]
            t_s += s
            t_r += r
            print(f"      {m:<9}{s:>13,.0f}{r:>12,.0f}{s + r:>13,.0f}")
        print(f"      {'합계':<9}{t_s:>13,.0f}{t_r:>12,.0f}{t_s + t_r:>13,.0f}  (부가세 포함)")


if __name__ == "__main__":
    main()
